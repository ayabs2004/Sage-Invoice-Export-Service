from flask import Flask, request, send_file, jsonify
import os
import sys
import pandas as pd
import pyodbc
from flask_cors import CORS
import io

def get_resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def get_logo_path(company):
    """Cherche le logo à côté de l'exe (png, jpeg, jpg)."""
    exe_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.abspath(".")
    for ext in ['png', 'jpeg', 'jpg']:
        candidate = os.path.join(exe_dir, "logos", f"{company}.{ext}")
        if os.path.exists(candidate):
            return candidate
    return None

def load_env_file(env_path=".env"):
    if os.path.exists(env_path):
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, value = line.split("=", 1)
                    os.environ[key.strip()] = value.strip()

load_env_file()

app = Flask(__name__)
CORS(app)

@app.route('/')
@app.route('/index.html')
def serve_index():
    return send_file(get_resource_path('index.html'))

@app.route('/generate-excel', methods=['POST'])
def generate_excel():
    data = request.json
    if not data:
        return jsonify({"error": "Aucune donnée JSON fournie"}), 400

    month = data.get("month")
    invoice_type = data.get("type")
    company = data.get("company")

    if not month or not invoice_type or not company:
        return jsonify({"error": "La société, le mois et le type sont obligatoires"}), 400

    db_env_key = f"DB_DATABASE_{company.upper()}"
    db_name = os.getenv(db_env_key)
    db_doc = os.getenv(f"{db_env_key}_DOC")
    db_tiers = os.getenv(f"{db_env_key}_TIERS")
    connect_db = db_name or db_doc or db_tiers

    if not connect_db:
        return jsonify({
            "error": f"La configuration de base de données pour '{company}' est manquante.",
            "details": f"Veuillez ajouter {db_env_key}=NomBase dans votre .env"
        }), 400

    type_mapping = {
        "vente": [7], "7": [7], 7: [7],
        "achat": [6], "6": [6], 6: [6],
        "achat_vente": [6, 7]
    }

    target_types = type_mapping.get(invoice_type)
    if not target_types:
        try:
            if isinstance(invoice_type, (int, float)):
                target_types = [int(invoice_type)]
            elif isinstance(invoice_type, str) and invoice_type.isdigit():
                target_types = [int(invoice_type)]
            else:
                return jsonify({"error": "Type de facture invalide"}), 400
        except:
            return jsonify({"error": "Type de facture invalide"}), 400

    company_up = company.upper()
    calculate_from_lines = True
    col_piece = "DO_Piece"
    table_doc = "F_DOCENTETE"

    db_doc = db_doc or connect_db
    db_tiers = db_tiers or connect_db
    db_ligne = os.getenv(f"DB_DATABASE_{company_up}_LIGNE", db_doc)
    table_tiers = "F_COMPTET"
    table_ligne = "F_DOCLIGNE"

    col_ligne_ht = "DL_MontantHT"
    col_ligne_ttc = "DL_MontantTTC"
    col_ht = "DO_TotalHT" if not calculate_from_lines else None
    col_ttc = "DO_TotalTTC" if not calculate_from_lines else None

    conn = None
    try:
        driver_name = os.getenv('DB_DRIVER', 'SQL Server')
        trusted_connection = os.getenv('DB_TRUSTED_CONNECTION', 'false').lower() in ['true', 'yes', '1']
        conn_str = f"DRIVER={{{driver_name}}};SERVER={os.getenv('DB_SERVER')};DATABASE={connect_db}"
        if trusted_connection:
            conn_str += ";Trusted_Connection=yes"
        else:
            conn_str += f";UID={os.getenv('DB_USER')};PWD={os.getenv('DB_PASSWORD')}"

        print(f"Connecting with: {conn_str.split('PWD=')[0]}PWD=***" if 'PWD=' in conn_str else f"Connecting with: {conn_str}")
        conn = pyodbc.connect(conn_str)

        cast_piece = f"CAST(D.[{col_piece}] AS NVARCHAR(100))"
        cond_vente = f"(CAST(D.DO_Domaine AS INT) = 0 AND {cast_piece} LIKE 'FV%')"
        cond_achat = f"(CAST(D.DO_Domaine AS INT) = 1 AND {cast_piece} LIKE 'FA%')"

        if invoice_type == 'vente':
            type_condition = cond_vente
        elif invoice_type == 'achat':
            type_condition = cond_achat
        else:
            type_condition = f"({cond_vente} OR {cond_achat})"

        date_filter = "CAST(LEFT(CONVERT(NVARCHAR(30), D.DO_Date, 120), 7) AS NVARCHAR(10)) = CAST(? AS NVARCHAR(10))"
        date_select = "CONVERT(NVARCHAR(10), D.DO_Date, 103)"
        from_clause = f"[{db_doc}]..[{table_doc}] D"
        join_clause = f"[{db_tiers}]..[{table_tiers}] C ON CAST(D.DO_Tiers AS NVARCHAR(MAX)) COLLATE DATABASE_DEFAULT = CAST(C.CT_Num AS NVARCHAR(MAX)) COLLATE DATABASE_DEFAULT"

        if col_ht and col_ttc:
            select_ht = f"D.[{col_ht}]"
            select_ttc = f"D.[{col_ttc}]"
            select_tva = f"(ISNULL(D.[{col_ttc}],0) - ISNULL(D.[{col_ht}],0) - 1)"
        else:
            ligne_path = f"[{db_ligne}]..[{table_ligne}]"
            piece_where = f"CAST(W.[{col_piece}] AS NVARCHAR(MAX)) COLLATE DATABASE_DEFAULT = CAST(D.[{col_piece}] AS NVARCHAR(MAX)) COLLATE DATABASE_DEFAULT"
            select_ht = f"(SELECT ISNULL(SUM(W.[{col_ligne_ht}]),0) FROM {ligne_path} W WHERE {piece_where})"
            select_ttc = f"(SELECT ISNULL(SUM(W.[{col_ligne_ttc}]),0) FROM {ligne_path} W WHERE {piece_where})"
            select_tva = f"( (SELECT ISNULL(SUM(W.[{col_ligne_ttc}] - W.[{col_ligne_ht}]),0) FROM {ligne_path} W WHERE {piece_where}) - 1 )"

        query = f"""
            SET ANSI_WARNINGS OFF;
            SELECT 
                {date_select} as DateFac, 
                CAST(D.[{col_piece}] AS NVARCHAR(MAX)) as Reference, 
                CAST(C.CT_Intitule AS NVARCHAR(MAX)) as Tiers, 
                {select_ht} as TotalHT,
                {select_tva} as TotalTVA,
                {select_ttc} as TotalTTC,
                CAST(D.DO_Type AS INT) as TypeDoc,
                CAST(D.DO_Domaine AS INT) as Domaine
            FROM {from_clause}
            LEFT JOIN {join_clause}
            WHERE {type_condition} 
            AND {date_filter}
            ORDER BY D.DO_Piece ASC
        """

        sql_params = [str(month)]
        import datetime
        now = datetime.datetime.now().strftime("%H:%M:%S")
        print(f"[{now}] --- DEBUG SQL ---")
        print(f"[{now}] Database: {connect_db}")

        try:
            inspect_cursor = conn.cursor()
            for t_name in [table_doc, table_ligne, table_tiers]:
                inspect_cursor.execute(f"SELECT COLUMN_NAME, DATA_TYPE FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = '{t_name}'")
                for col_info in inspect_cursor.fetchall():
                    if 'ntext' in str(col_info).lower() or 'text' in str(col_info).lower():
                        print(f"  [!!!] Column {col_info[0]} is type {col_info[1]}")
        except Exception as insp_e:
            print(f"[{now}] Schema inspection failed: {str(insp_e)}")

        print(f"[{now}] Query: {query}")
        print(f"[{now}] Params: {sql_params}")

        cursor = conn.cursor()
        try:
            cursor.execute(query, sql_params)
        except Exception as exec_e:
            print(f"[{now}] Execution failed: {str(exec_e)}")
            raise exec_e

        columns = [column[0] for column in cursor.description]
        rows = cursor.fetchall()
        df = pd.DataFrame.from_records([list(r) for r in rows], columns=columns)
        df.columns = ["Date", "Référence", "Tiers", "H.T", "TVA", "TTC", "DO_Type", "Domaine"]

        df["DO_Type"] = pd.to_numeric(df["DO_Type"], errors="coerce").fillna(0).astype(int)
        df["Domaine"] = pd.to_numeric(df["Domaine"], errors="coerce").fillna(0).astype(int)
        df["Référence"] = df["Référence"].fillna("").astype(str)

        print(f"[{now}] Sample data (Top 5):")
        print(df[["Référence", "DO_Type"]].head(5).to_string(index=False))
        print(f"[{now}] Unique types found: {df['DO_Type'].unique().tolist()}")

        if df.empty:
            return jsonify({"error": "Aucune donnée trouvée pour ces critères"}), 404

        try:
            y, m = month.split('-')
            months_fr = {
                "01": "Janvier", "02": "Février", "03": "Mars", "04": "Avril",
                "05": "Mai", "06": "Juin", "07": "Juillet", "08": "Août",
                "09": "Septembre", "10": "Octobre", "11": "Novembre", "12": "Décembre"
            }
            month_label = f"{months_fr.get(m, m)} {y}"
        except:
            month_label = month

        output_file_name = f"sage_export_{invoice_type}_{month}.xlsx"
        output_buffer = io.BytesIO()

        print(f"Generating file in memory: {output_file_name}")

        with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
            workbook = writer.book
            worksheet = workbook.create_sheet('Export')
            writer.sheets['Export'] = worksheet

            workbook.calculation.calcMode = 'auto'
            if hasattr(workbook.properties, 'calcId'):
                workbook.properties.calcId = None

            from openpyxl.styles import Font, PatternFill, Alignment, Side, Border
            from openpyxl.utils import get_column_letter, column_index_from_string
            from openpyxl.drawing.image import Image as XLImage

            header_fill = PatternFill(start_color="F2F1E9", end_color="F2F1E9", fill_type="solid")
            header_font = Font(name='Calibri', size=11, bold=True, color="000000")
            title_font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
            sub_title_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            sub_title_font = Font(name='Calibri', size=12, bold=True, color="000000")
            total_fill = PatternFill(start_color="F8C89A", end_color="F8C89A", fill_type="solid")
            total_font = Font(name='Calibri', size=11, bold=True)
            main_title_fill = PatternFill(start_color="4FB3C8", end_color="4FB3C8", fill_type="solid")
            border_style = Side(style='thin', color="000000")
            thin_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

            def write_table(dataframe, start_col_idx, table_title, tiers_label):
                cell_sub = worksheet.cell(row=3, column=start_col_idx, value=table_title)
                cell_sub.font = sub_title_font
                cell_sub.fill = sub_title_fill
                cell_sub.alignment = Alignment(horizontal='center')
                worksheet.merge_cells(start_row=3, start_column=start_col_idx, end_row=3, end_column=start_col_idx + 5)

                headers = ["Date", "Référence", tiers_label, "H.T", "TVA", "TTC"]
                for i, h in enumerate(headers):
                    c = worksheet.cell(row=4, column=start_col_idx + i, value=h)
                    c.fill = header_fill
                    c.font = header_font
                    c.alignment = Alignment(horizontal='center')
                    c.border = thin_border

                for row_counter, (r_idx, row) in enumerate(dataframe.iterrows()):
                    row_vals = [row["Date"], row["Référence"], row["Tiers"], row["H.T"], row["TVA"], row["TTC"]]
                    for c_idx, val in enumerate(row_vals):
                        c = worksheet.cell(row=5 + row_counter, column=start_col_idx + c_idx, value=val)
                        c.border = thin_border
                        if c_idx >= 3:
                            c.number_format = '#,##0.00'

                last_row = 5 + len(dataframe)
                tot_lbl = worksheet.cell(row=last_row, column=start_col_idx + 2, value="TOTAL")
                tot_lbl.fill = total_fill
                tot_lbl.font = total_font
                tot_lbl.alignment = Alignment(horizontal='right')
                tot_lbl.border = thin_border

                for i in range(3):
                    col_letter = get_column_letter(start_col_idx + 3 + i)
                    formula = f"=SUM({col_letter}5:{col_letter}{last_row - 1})"
                    c = worksheet.cell(row=last_row, column=start_col_idx + 3 + i, value=formula)
                    c.fill = total_fill
                    c.font = total_font
                    c.alignment = Alignment(horizontal='right')
                    c.border = thin_border
                    c.number_format = '#,##0.00'

            # ── ROW 1 : Logo (A1) + Titre (A1 mergée → dernière colonne) ────────
            worksheet.row_dimensions[1].height = 55
            worksheet.column_dimensions['A'].width = 22

            title_text = f"DECLARATION {month_label.upper()}"

            if len(target_types) > 1:
                end_col_letter = 'M'
            else:
                end_col_letter = 'F'

            end_col = column_index_from_string(end_col_letter)

            # Merge unique sur toute la ligne 1
            worksheet.merge_cells(f'A1:{end_col_letter}1')

            # Fond bleu sur toute la ligne 1
            for col_idx in range(1, end_col + 1):
                worksheet.cell(row=1, column=col_idx).fill = main_title_fill

            # Titre dans la cellule maître A1
            title_cell = worksheet['A1']
            title_cell.value = title_text
            title_cell.font = title_font
            title_cell.fill = main_title_fill
            title_cell.alignment = Alignment(horizontal='right', vertical='center')

            # Logo flottant par-dessus A1
            logo_path = get_logo_path(company)
            if logo_path:
                img = XLImage(logo_path)
                img.width = 150
                img.height = 50
                worksheet.add_image(img, "A1")
            else:
                print(f"[WARN] Logo non trouvé pour : {company}")

            # ── Tableaux ─────────────────────────────────────────────────────────
            if len(target_types) > 1:
                df_vente = df[(df['Domaine'] == 0) & (df['Référence'].str.startswith('FV'))]
                df_achat = df[(df['Domaine'] == 1) & (df['Référence'].str.startswith('FA'))]

                write_table(df_vente, 1, "VENTE", "Client")
                write_table(df_achat, 8, "ACHAT", "Fournisseur")

                for col_idx in list(range(1, 7)) + list(range(8, 14)):
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = 18
            else:
                is_vente = (invoice_type == 'vente' or 7 in target_types)
                t_label = "Client" if is_vente else "Fournisseur"
                t_title = "VENTE" if is_vente else "ACHAT"

                write_table(df, 1, t_title, t_label)

                for col in ['B', 'C', 'D', 'E', 'F']:
                    worksheet.column_dimensions[col].width = 20

        output_buffer.seek(0)
        return send_file(
            output_buffer,
            as_attachment=True,
            download_name=output_file_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        import traceback
        err_msg = f"Erreur serveur: {str(e)}\n{traceback.format_exc()}"
        print(f"--- ERROR ---\n{err_msg}")
        return jsonify({"error": str(e), "details": traceback.format_exc()}), 500
    finally:
        if conn:
            conn.close()


@app.route('/inspect', methods=['GET'])
def inspect_table():
    company = request.args.get("company", "jacob")
    table = request.args.get("table", "F_DOCENTETE")

    db_env_key = f"DB_DATABASE_{company.upper()}"
    db_name = os.getenv(db_env_key) or os.getenv(f"{db_env_key}_DOC")

    if not db_name:
        return jsonify({"error": "Base non configurée pour cette société"}), 400

    conn = None
    try:
        driver_name = os.getenv('DB_DRIVER', 'SQL Server')
        conn_str = f"DRIVER={{{driver_name}}};SERVER={os.getenv('DB_SERVER')};DATABASE={db_name}"
        trusted_connection = os.getenv('DB_TRUSTED_CONNECTION', 'false').lower() in ['true', 'yes', '1']
        if trusted_connection:
            conn_str += ";Trusted_Connection=yes"
        else:
            conn_str += f";UID={os.getenv('DB_USER')};PWD={os.getenv('DB_PASSWORD')}"

        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        cursor.execute(f"SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = '{table}'")
        columns = [row[0] for row in cursor.fetchall()]

        if not columns:
            cursor.execute(f"SELECT name FROM sys.columns WHERE object_id = OBJECT_ID('{table}')")
            columns = [row[0] for row in cursor.fetchall()]

        return jsonify({"company": company, "table": table, "database": db_name, "columns": columns})

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8000, debug=True)